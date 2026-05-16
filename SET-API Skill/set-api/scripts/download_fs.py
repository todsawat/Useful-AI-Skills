#!/usr/bin/env python3
"""
SET (Stock Exchange of Thailand) news & financial statement tool.
No browser required — uses urllib + Node.js (for __NUXT__ eval) + zipfile.

Modes:
  news       — Search and display SET company news
  financial  — Download financial statement files (ZIP/PDF)
  quick      — Quick view F45 summary (สรุปผลการดำเนินงาน) without downloading
  stock      — Stock information (price, historical trading, shareholders, etc.)
  warrant    — Warrant information (related warrants, exercise price, maturity, etc.)
  report     — Annual reports, 56-1 One Report, ESG reports (list or download)

Usage:
  # News for a single stock
  python3 download_fs.py news --symbol BH

  # News for all stocks on a date
  python3 download_fs.py news --from-date 14/05/2026 --to-date 14/05/2026

  # Quick view F45 earnings summary (no file download)
  python3 download_fs.py quick --symbol BH --quarters 4

  # Quick view all F45 filed today
  python3 download_fs.py quick --from-date 14/05/2026 --to-date 14/05/2026

  # Financial statements for a stock (8 quarters)
  python3 download_fs.py financial --symbol BH --quarters 8 --out ./output

  # Financial statements for ALL stocks on a specific date
  python3 download_fs.py financial --from-date 14/05/2026 --to-date 14/05/2026 --out ./output

  # List only (no download)
  python3 download_fs.py financial --symbol BH --quarters 4 --list-only

  # Stock info — all sections
  python3 download_fs.py stock --symbol BH

  # Stock info — specific sections only
  python3 download_fs.py stock --symbol BH --sections price,shareholders

  # Historical trading with date range
  python3 download_fs.py stock --symbol BH --sections historical --from-date 01/01/2026 --to-date 14/05/2026

Environment:
  Requires Node.js (for evaluating SET's __NUXT__ SSR data).
"""

import argparse
import http.cookiejar
import json
import os
import re
import subprocess
import sys
import tempfile
import time
import urllib.request
import zipfile
from datetime import datetime, timedelta

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
API_HEADERS = {
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.set.or.th/th/market/news-and-alert/news",
}


# ── Shared helpers ──────────────────────────────────────────────────────────

def make_opener():
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    return opener, cj


def get_set_cookies(opener):
    """Visit SET news page to obtain Incapsula session cookies."""
    req = urllib.request.Request(
        "https://www.set.or.th/th/market/news-and-alert/news",
        headers={"User-Agent": UA, "Accept": "text/html"},
    )
    opener.open(req, timeout=30)


def get_weblink_cookies(opener):
    """Visit weblink.set.or.th for Incapsula cookies (needed for file downloads)."""
    try:
        req = urllib.request.Request(
            "https://weblink.set.or.th/", headers={"User-Agent": UA, "Accept": "text/html"}
        )
        opener.open(req, timeout=15)
    except Exception:
        pass


def search_news(opener, symbol=None, from_date=None, to_date=None,
                per_page=None, page=1, keyword=None):
    """
    Call SET News Search API.
    Returns (total_count, news_list).
    Date format: DD/MM/YYYY.
    """
    params = ["sourceId=company", "securityType=S", "lang=th"]
    if symbol:
        params.append(f"symbol={symbol}")
    if from_date:
        params.append(f"fromDate={from_date}")
    if to_date:
        params.append(f"toDate={to_date}")
    if per_page:
        params.append(f"perPage={per_page}")
    if page and page > 1:
        params.append(f"page={page}")
    if keyword:
        params.append(f"keyword={urllib.request.quote(keyword)}")

    url = "https://www.set.or.th/api/set/news/search?" + "&".join(params)
    req = urllib.request.Request(url, headers=API_HEADERS)
    data = json.loads(opener.open(req, timeout=30).read().decode("utf-8"))
    return data.get("totalCount", 0), data.get("newsInfoList", [])


def search_all_news_paginated(opener, symbol=None, from_date=None, to_date=None,
                              keyword=None, max_results=None):
    """
    Fetch all news with pagination. Returns list of news items.
    """
    all_news = []
    page = 1
    per_page = 200  # max reasonable page size

    while True:
        total, news = search_news(
            opener, symbol=symbol, from_date=from_date, to_date=to_date,
            per_page=per_page, page=page, keyword=keyword,
        )
        if not news:
            break
        all_news.extend(news)
        if max_results and len(all_news) >= max_results:
            all_news = all_news[:max_results]
            break
        if len(all_news) >= total:
            break
        page += 1
        time.sleep(0.2)

    return all_news


def get_download_url(opener, news_id, symbol):
    """
    Fetch newsdetails page, eval __NUXT__ via Node.js, extract downloadUrl.
    Returns (headline, download_url) or (None, None).
    """
    detail_url = (
        f"https://www.set.or.th/th/market/news-and-alert/newsdetails"
        f"?id={news_id}&symbol={symbol}"
    )
    req = urllib.request.Request(
        detail_url, headers={"User-Agent": UA, "Accept": "text/html"}
    )
    html = opener.open(req, timeout=30).read().decode("utf-8")

    match = re.search(r"window\.__NUXT__\s*=\s*(.+?);\s*</script>", html, re.DOTALL)
    if not match:
        return None, None

    with tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False) as f:
        f.write(match.group(1))
        tmp_js = f.name

    node_code = f"""
    const fs = require('fs');
    const d = eval(fs.readFileSync('{tmp_js}','utf-8'));
    const nd = d.state && d.state.news && d.state.news.newsDetails;
    if (nd && nd.downloadUrl) {{
        console.log(JSON.stringify({{headline: nd.headline, url: nd.downloadUrl}}));
    }}
    """
    try:
        result = subprocess.run(
            ["node", "-e", node_code], capture_output=True, text=True, timeout=10
        )
        if result.stdout.strip():
            info = json.loads(result.stdout.strip())
            return info["headline"], info["url"]
    finally:
        os.unlink(tmp_js)

    return None, None


def safe_folder_name(headline, symbol):
    """Create a filesystem-safe folder name from headline."""
    name = re.sub(r"[^\w\s/]", "", headline).strip().replace(" ", "_").replace("/", "-")
    name = f"{symbol}_{name}"
    while len(name.encode("utf-8")) > 200:
        name = name[:-1]
    return name.rstrip("_")


def download_file(opener, download_url, headline, symbol, out_dir):
    """
    Download a file (ZIP or PDF) from weblink.set.or.th.
    ZIPs are extracted; PDFs are saved as-is.
    Returns (path, file_list) or (None, None).
    """
    folder_name = safe_folder_name(headline, symbol)

    req = urllib.request.Request(
        download_url, headers={"User-Agent": UA, "Referer": "https://www.set.or.th/"}
    )
    resp = opener.open(req, timeout=60)
    content = resp.read()

    if content[:2] == b"PK":
        zip_path = os.path.join(out_dir, f"{folder_name}.zip")
        extract_dir = os.path.join(out_dir, folder_name)
        with open(zip_path, "wb") as f:
            f.write(content)
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(extract_dir)
            files = zf.namelist()
        return extract_dir, files

    if content[:5] == b"%PDF-":
        pdf_path = os.path.join(out_dir, f"{folder_name}.pdf")
        with open(pdf_path, "wb") as f:
            f.write(content)
        return pdf_path, [f"{folder_name}.pdf"]

    return None, None


def get_news_content(opener, news_id, symbol):
    """
    Fetch newsdetails page, eval __NUXT__ via Node.js, extract content text.
    Returns (headline, content, downloadUrl) or (None, None, None).
    """
    detail_url = (
        f"https://www.set.or.th/th/market/news-and-alert/newsdetails"
        f"?id={news_id}&symbol={symbol}"
    )
    req = urllib.request.Request(
        detail_url, headers={"User-Agent": UA, "Accept": "text/html"}
    )
    html = opener.open(req, timeout=30).read().decode("utf-8")

    match = re.search(r"window\.__NUXT__\s*=\s*(.+?);\s*</script>", html, re.DOTALL)
    if not match:
        return None, None, None

    with tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False) as f:
        f.write(match.group(1))
        tmp_js = f.name

    node_code = f"""
    const fs = require('fs');
    const d = eval(fs.readFileSync('{tmp_js}','utf-8'));
    const nd = d.state && d.state.news && d.state.news.newsDetails;
    if (nd) {{
        console.log(JSON.stringify({{
            headline: nd.headline || '',
            content: nd.content || '',
            downloadUrl: nd.downloadUrl || ''
        }}));
    }}
    """
    try:
        result = subprocess.run(
            ["node", "-e", node_code], capture_output=True, text=True, timeout=10
        )
        if result.stdout.strip():
            info = json.loads(result.stdout.strip())
            return info["headline"], info["content"], info["downloadUrl"]
    finally:
        os.unlink(tmp_js)

    return None, None, None


def strip_disclaimer(content):
    """Remove the SET standard disclaimer from the end of content."""
    disc_idx = content.find("______")
    if disc_idx > 0:
        content = content[:disc_idx]
    # Also strip HTML tags
    content = re.sub(r"<[^>]+>", "", content)
    return content.strip()


# ── Mode: news ──────────────────────────────────────────────────────────────

def cmd_news(args):
    """Search and display SET company news."""
    opener, _ = make_opener()

    print("[1/2] Getting session cookies ...")
    get_set_cookies(opener)

    print("[2/2] Searching news ...")
    news_list = search_all_news_paginated(
        opener,
        symbol=args.symbol.upper() if args.symbol else None,
        from_date=args.from_date,
        to_date=args.to_date,
        keyword=args.keyword,
        max_results=args.limit,
    )

    if not news_list:
        print("  No news found.")
        return

    # Optional: filter by headline keyword
    if args.filter:
        pat = re.compile(args.filter, re.IGNORECASE)
        news_list = [n for n in news_list if pat.search(n.get("headline", ""))]

    print(f"\nFound {len(news_list)} news items:\n")

    if args.format == "json":
        print(json.dumps(news_list, ensure_ascii=False, indent=2))
    else:
        for n in news_list:
            dt = n.get("datetime", "")[:16].replace("T", " ")
            sym = n.get("symbol", "???")
            hl = n.get("headline", "")
            nid = n.get("id", "")
            print(f"  {dt} | {sym:8s} | {hl}")
            if args.show_id:
                print(f"{'':19s} | id={nid}")

    print(f"\nTotal: {len(news_list)} items")


# ── Mode: quick ────────────────────────────────────────────────────────────

def cmd_quick(args):
    """Quick view F45 summaries (สรุปผลการดำเนินงาน) — no file downloads."""
    opener, _ = make_opener()

    print("[1/3] Getting session cookies ...")
    get_set_cookies(opener)

    # Determine search params
    symbol = args.symbol.upper() if args.symbol else None
    from_date = args.from_date
    to_date = args.to_date

    if not from_date and symbol:
        today = datetime.now()
        from_date = (today - timedelta(days=365 * 3)).strftime("%d/%m/%Y")
        to_date = today.strftime("%d/%m/%Y")
    elif not from_date and not symbol:
        today = datetime.now()
        from_date = today.strftime("%d/%m/%Y")
        to_date = from_date

    desc = symbol if symbol else "all stocks"
    print(f"[2/3] Searching F45 summaries: {desc} ({from_date} - {to_date}) ...")

    all_news = search_all_news_paginated(
        opener, symbol=symbol, from_date=from_date, to_date=to_date
    )

    # Filter: "สรุปผลการดำเนินงาน" (F45 forms only, not "เผยแพร่งบ..." press releases)
    f45_news = [n for n in all_news
                if "สรุปผลการดำเนินงาน" in n.get("headline", "")
                and "F45" in n.get("headline", "")]

    # Deduplicate
    seen = {}
    for n in f45_news:
        sym = n.get("symbol", "")
        key = sym + "|" + re.sub(r"\s*\(แก้ไข\)", "", n["headline"])
        if key not in seen:
            seen[key] = n
    unique = list(seen.values())

    # Limit
    if args.quarters and symbol:
        unique = unique[: args.quarters]
    elif args.limit:
        unique = unique[: args.limit]

    if not unique:
        print("  No F45 summaries found")
        sys.exit(1)

    symbols_found = sorted(set(n["symbol"] for n in unique))
    print(f"  Found {len(unique)} F45 items across {len(symbols_found)} symbols")

    # Step 3: Fetch content from each newsdetails page
    print(f"[3/3] Fetching F45 content ...")
    results = []

    for i, n in enumerate(unique):
        sym = n.get("symbol", "???")
        headline, content, dl_url = get_news_content(opener, n["id"], sym)
        if not content:
            print(f"  [{i+1}/{len(unique)}] {sym}: no content, skipping")
            continue

        results.append({
            "symbol": sym,
            "headline": headline or n["headline"],
            "date": n["datetime"][:10],
            "content": strip_disclaimer(content),
        })

        time.sleep(0.3)

    if not results:
        print("  No F45 content found")
        sys.exit(1)

    # Output — always raw content for AI to interpret
    if args.format == "json":
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        for r in results:
            print(f"\n{'='*70}")
            print(f"  {r['symbol']} | {r['date']} | {r['headline']}")
            print(f"{'='*70}")
            print(r["content"])
        print(f"\n--- Total: {len(results)} items ---")

    return results


# ── Mode: financial ─────────────────────────────────────────────────────────

def cmd_financial(args):
    """Download financial statement files."""
    opener, _ = make_opener()
    out_dir = args.out
    os.makedirs(out_dir, exist_ok=True)

    print("[1/5] Getting session cookies ...")
    get_set_cookies(opener)

    # Determine search params
    symbol = args.symbol.upper() if args.symbol else None
    from_date = args.from_date
    to_date = args.to_date

    if not from_date and symbol:
        # Default: search last N years for a single stock
        today = datetime.now()
        from_date = (today - timedelta(days=365 * 3)).strftime("%d/%m/%Y")
        to_date = today.strftime("%d/%m/%Y")
    elif not from_date and not symbol:
        # All stocks without date = today only
        today = datetime.now()
        from_date = today.strftime("%d/%m/%Y")
        to_date = from_date

    desc = symbol if symbol else "all stocks"
    print(f"[2/5] Searching financial statements: {desc} ({from_date} - {to_date}) ...")

    all_news = search_all_news_paginated(
        opener, symbol=symbol, from_date=from_date, to_date=to_date
    )

    # Filter: only "งบการเงิน" headlines
    fs_news = [n for n in all_news if "งบการเงิน" in n.get("headline", "")]

    # Deduplicate: per symbol+quarter, keep first (newest = corrected version if exists)
    seen = {}
    for n in fs_news:
        sym = n.get("symbol", "")
        key = sym + "|" + re.sub(r"\s*\(แก้ไข\)", "", n["headline"])
        if key not in seen:
            seen[key] = n

    unique = list(seen.values())

    # Limit
    if args.quarters and symbol:
        unique = unique[: args.quarters]
    elif args.limit:
        unique = unique[: args.limit]

    if not unique:
        print(f"  No financial statements found")
        sys.exit(1)

    symbols_found = sorted(set(n["symbol"] for n in unique))
    print(f"  Found {len(unique)} items across {len(symbols_found)} symbols: {', '.join(symbols_found[:10])}{'...' if len(symbols_found) > 10 else ''}")

    # Step 3: Get download URLs
    print(f"[3/5] Extracting download URLs ...")
    downloads = []
    for i, n in enumerate(unique):
        sym = n.get("symbol", "???")
        headline, url = get_download_url(opener, n["id"], sym)
        if headline and url:
            downloads.append({
                "headline": headline, "url": url,
                "date": n["datetime"][:10], "symbol": sym,
            })
            print(f"  [{i+1}/{len(unique)}] {sym}: {headline}")
        else:
            print(f"  [{i+1}/{len(unique)}] {sym}: {n['headline']}: no URL, skipping")
        time.sleep(0.3)

    if not downloads:
        print("  No download URLs found")
        sys.exit(1)

    if args.list_only:
        print(f"\nAvailable downloads ({len(downloads)}):")
        for d in downloads:
            print(f"  {d['date']} | {d['symbol']:8s} | {d['headline']}")
            print(f"{'':13s} {d['url']}")
        sys.exit(0)

    # Step 4: weblink cookies
    print("[4/5] Getting download cookies ...")
    get_weblink_cookies(opener)

    # Step 5: Download
    print(f"[5/5] Downloading {len(downloads)} files to {os.path.abspath(out_dir)} ...")
    results = []
    for i, dl in enumerate(downloads):
        print(f"  [{i+1}/{len(downloads)}] {dl['symbol']}: {dl['headline']}")
        try:
            path, files = download_file(
                opener, dl["url"], dl["headline"], dl["symbol"], out_dir
            )
            if path:
                print(f"    -> {len(files)} files: {os.path.basename(path)}")
                results.append({"symbol": dl["symbol"], "headline": dl["headline"],
                                "path": path, "files": files})
                if args.read_content and os.path.isdir(path):
                    for fname in sorted(os.listdir(path)):
                        fpath = os.path.join(path, fname)
                        content = read_file_content(fpath)
                        if content:
                            print(f"\n{'='*70}")
                            print(f"FILE: {fname}  [{dl['symbol']} | {dl['headline']}]")
                            print(f"{'='*70}")
                            print(content)
            else:
                print(f"    -> Failed (unknown format)")
        except Exception as e:
            print(f"    -> Error: {e}")
        time.sleep(0.5)

    # Summary
    print(f"\nDone! Downloaded {len(results)}/{len(downloads)} items")
    for r in results:
        print(f"  [{r['symbol']}] {r['headline']}")
        for f in r["files"]:
            print(f"    - {f}")

    return results


# ── File content readers ────────────────────────────────────────────────────

def read_file_content(path):
    """Read text content from xlsx, xls, docx, or doc file. Returns string or None."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            lines = []
            for name in wb.sheetnames:
                ws = wb[name]
                lines.append(f"\n[Sheet: {name}]")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append("\t".join(cells))
            return "\n".join(lines)
        except ImportError:
            return "[openpyxl not installed — pip install openpyxl]"
        except Exception as e:
            return f"[xlsx read error: {e}]"

    elif ext == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            lines = []
            for sheet in wb.sheets():
                lines.append(f"\n[Sheet: {sheet.name}]")
                for r in range(sheet.nrows):
                    cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    if any(c.strip() for c in cells):
                        lines.append("\t".join(cells))
            return "\n".join(lines)
        except ImportError:
            return "[xlrd not installed — pip install xlrd]"
        except Exception as e:
            return f"[xls read error: {e}]"

    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return "[python-docx not installed — pip install python-docx]"
        except Exception as e:
            return f"[docx read error: {e}]"

    elif ext == ".doc":
        # Try external tools in order: textutil (macOS), antiword, LibreOffice
        try:
            result = subprocess.run(
                ["textutil", "-convert", "txt", "-stdout", path],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout
        except FileNotFoundError:
            pass

        try:
            result = subprocess.run(
                ["antiword", path],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout
        except FileNotFoundError:
            pass

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                subprocess.run(
                    ["soffice", "--headless", "--convert-to", "txt:Text",
                     "--outdir", tmpdir, path],
                    capture_output=True, timeout=60,
                )
                txt_path = os.path.join(
                    tmpdir, os.path.splitext(os.path.basename(path))[0] + ".txt"
                )
                if os.path.exists(txt_path):
                    with open(txt_path, encoding="utf-8", errors="replace") as f:
                        return f.read()
        except FileNotFoundError:
            pass

        return "[.doc: ต้องการ textutil (macOS built-in), antiword (brew install antiword), หรือ LibreOffice]"

    return None


# ── Mode: stock ────────────────────────────────────────────────────────────

# Available stock API sections and their endpoints
STOCK_SECTIONS = {
    "price": {
        "label": "Price & Quote",
        "endpoints": [
            "/api/set/stock/{symbol}/info",
            "/api/set/stock/{symbol}/highlight-data",
            "/api/set/stock/{symbol}/price-performance",
        ],
    },
    "historical": {
        "label": "Historical Trading",
        "endpoints": [
            "/api/set/stock/{symbol}/historical-trading",
        ],
    },
    "rights": {
        "label": "Rights & Benefits (Corporate Actions)",
        "endpoints": [
            "/api/set/stock/{symbol}/corporate-action",
        ],
    },
    "shareholders": {
        "label": "Major Shareholders",
        "endpoints": [
            "/api/set/stock/{symbol}/shareholder",
        ],
    },
    "nvdr": {
        "label": "NVDR Holdings",
        "endpoints": [
            "/api/set/stock/{symbol}/nvdr-holder",
        ],
    },
    "profile": {
        "label": "Company Profile & Financials",
        "endpoints": [
            "/api/set/stock/{symbol}/profile",
            "/api/set/stock/{symbol}/key-financial-data",
        ],
    },
}


def fetch_stock_api(opener, url):
    """Fetch a single SET stock API endpoint. Returns parsed JSON or error string."""
    req = urllib.request.Request(url, headers=API_HEADERS)
    try:
        resp = opener.open(req, timeout=30)
        return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        return {"_error": f"HTTP {e.code}", "_url": url}
    except Exception as e:
        return {"_error": str(e), "_url": url}


def cmd_stock(args):
    """Fetch stock information from SET APIs."""
    symbol = args.symbol.upper()
    opener, _ = make_opener()

    print(f"[1/2] Getting session cookies ...", file=sys.stderr)
    get_set_cookies(opener)

    # Determine which sections to fetch
    if args.sections:
        requested = [s.strip().lower() for s in args.sections.split(",")]
        invalid = [s for s in requested if s not in STOCK_SECTIONS]
        if invalid:
            print(f"Error: unknown section(s): {', '.join(invalid)}", file=sys.stderr)
            print(f"Available: {', '.join(STOCK_SECTIONS.keys())}", file=sys.stderr)
            sys.exit(1)
        sections = {k: STOCK_SECTIONS[k] for k in requested}
    else:
        sections = STOCK_SECTIONS

    print(f"[2/2] Fetching {symbol} data: {', '.join(sections.keys())} ...", file=sys.stderr)

    result = {"symbol": symbol, "sections": {}}

    for section_key, section_def in sections.items():
        section_data = {}
        for ep_template in section_def["endpoints"]:
            ep = ep_template.replace("{symbol}", symbol)
            url = f"https://www.set.or.th{ep}"

            # For historical trading, add date params if provided
            if "historical-trading" in ep and (args.from_date or args.to_date):
                params = []
                if args.from_date:
                    params.append(f"fromDate={args.from_date}")
                if args.to_date:
                    params.append(f"toDate={args.to_date}")
                url += "?" + "&".join(params)

            api_name = ep.split("/")[-1]
            data = fetch_stock_api(opener, url)
            section_data[api_name] = data

        result["sections"][section_key] = {
            "label": section_def["label"],
            "data": section_data,
        }

    # Output — always raw JSON for AI to interpret
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


# ── Mode: warrant ──────────────────────────────────────────────────────────

def cmd_warrant(args):
    """Fetch warrant information for a stock."""
    symbol = args.symbol.upper()
    opener, _ = make_opener()

    print(f"[1/3] Getting session cookies ...", file=sys.stderr)
    get_set_cookies(opener)

    result = {"symbol": symbol, "warrants": []}

    # Step 1: Get related warrant products
    print(f"[2/3] Checking warrants for {symbol} ...", file=sys.stderr)
    url = f"https://www.set.or.th/api/set/stock/{symbol}/related-product/W"
    related = fetch_stock_api(opener, url)

    warrant_list = related.get("relatedProducts", [])
    if not warrant_list:
        result["message"] = f"No warrants found for {symbol}"
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return result

    print(f"  Found {len(warrant_list)} warrant(s)", file=sys.stderr)

    # Step 2: Get detailed profile for each warrant
    print(f"[3/3] Fetching warrant details ...", file=sys.stderr)
    for w in warrant_list:
        w_symbol = w.get("symbol", "")
        if not w_symbol:
            continue

        # Get warrant profile (has exercise price, ratio, maturity, etc.)
        profile_url = f"https://www.set.or.th/api/set/stock/{w_symbol}/profile"
        profile = fetch_stock_api(opener, profile_url)

        result["warrants"].append({
            "quote": w,         # realtime price data
            "profile": profile,  # warrant terms & details
        })
        time.sleep(0.2)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


# ── Mode: report ───────────────────────────────────────────────────────────

REPORT_TYPES = {
    "one": {
        "label": "56-1 One Report (แบบ 56-1 One Report)",
        "endpoint": "/api/set/company/{symbol}/report/one",
    },
    "annual": {
        "label": "Annual Report (รายงานประจำปี)",
        "endpoint": "/api/set/company/{symbol}/report/annual",
    },
    "form56": {
        "label": "Form 56-1 (แบบ 56-1)",
        "endpoint": "/api/set/company/{symbol}/report/form56",
    },
    "esg": {
        "label": "ESG Report",
        "endpoint": "/api/set/company/{symbol}/report/esg",
    },
}


def cmd_report(args):
    """List or download annual reports / 56-1 One Report."""
    symbol = args.symbol.upper()
    opener, _ = make_opener()

    print(f"[1/3] Getting session cookies ...", file=sys.stderr)
    get_set_cookies(opener)

    # Determine which report types to fetch
    if args.type:
        requested = [t.strip().lower() for t in args.type.split(",")]
        invalid = [t for t in requested if t not in REPORT_TYPES]
        if invalid:
            print(f"Error: unknown report type(s): {', '.join(invalid)}", file=sys.stderr)
            print(f"Available: {', '.join(REPORT_TYPES.keys())}", file=sys.stderr)
            sys.exit(1)
        types = {k: REPORT_TYPES[k] for k in requested}
    else:
        types = REPORT_TYPES

    # Step 2: Fetch report listings
    print(f"[2/3] Fetching {symbol} reports: {', '.join(types.keys())} ...", file=sys.stderr)
    result = {"symbol": symbol, "reports": {}}

    for type_key, type_def in types.items():
        ep = type_def["endpoint"].replace("{symbol}", symbol)
        url = f"https://www.set.or.th{ep}"
        data = fetch_stock_api(opener, url)
        result["reports"][type_key] = {
            "label": type_def["label"],
            "items": data if isinstance(data, list) else [data],
        }

    # Also fetch board of directors (useful company info)
    if not args.type or "profile" in (args.type or ""):
        board_url = f"https://www.set.or.th/api/set/company/{symbol}/board-of-director"
        board = fetch_stock_api(opener, board_url)
        if isinstance(board, list) and board:
            result["board_of_directors"] = board

    if args.list_only or not args.out:
        # List mode — just show available reports
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return result

    # Step 3: Download reports
    out_dir = args.out
    os.makedirs(out_dir, exist_ok=True)
    print(f"[3/3] Downloading reports to {os.path.abspath(out_dir)} ...", file=sys.stderr)

    get_weblink_cookies(opener)

    downloaded = []
    for type_key, report_data in result["reports"].items():
        items = report_data["items"]
        if not isinstance(items, list):
            continue

        # Apply year filter
        if args.year:
            items = [i for i in items if i.get("year") == args.year]

        # Apply limit
        limit = args.limit or 1  # default: download latest only
        items = items[:limit]

        for item in items:
            dl_url = item.get("url", "")
            year = item.get("year", "?")
            if not dl_url:
                continue

            filename = f"{symbol}_{type_key}_{year}"
            print(f"  {type_key} {year}: {dl_url}", file=sys.stderr)

            req = urllib.request.Request(
                dl_url, headers={"User-Agent": UA, "Referer": "https://www.set.or.th/"}
            )
            try:
                resp = opener.open(req, timeout=60)
                content = resp.read()

                if content[:2] == b"PK":
                    zip_path = os.path.join(out_dir, f"{filename}.zip")
                    extract_dir = os.path.join(out_dir, filename)
                    with open(zip_path, "wb") as f:
                        f.write(content)
                    os.makedirs(extract_dir, exist_ok=True)
                    with zipfile.ZipFile(zip_path) as zf:
                        zf.extractall(extract_dir)
                        files = zf.namelist()
                    downloaded.append({"type": type_key, "year": year, "path": extract_dir, "files": files})
                    print(f"    -> {len(files)} files extracted", file=sys.stderr)
                elif content[:5] == b"%PDF-":
                    pdf_path = os.path.join(out_dir, f"{filename}.pdf")
                    with open(pdf_path, "wb") as f:
                        f.write(content)
                    downloaded.append({"type": type_key, "year": year, "path": pdf_path, "files": [f"{filename}.pdf"]})
                    print(f"    -> PDF saved", file=sys.stderr)
                else:
                    print(f"    -> Unknown format ({len(content)} bytes)", file=sys.stderr)
            except Exception as e:
                print(f"    -> Error: {e}", file=sys.stderr)

            time.sleep(0.5)

    result["downloaded"] = downloaded
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


# ── CLI ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="SET news & financial statement tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Search BH news
  %(prog)s news --symbol BH

  # All company news on a date
  %(prog)s news --from-date 14/05/2026 --to-date 14/05/2026

  # News with keyword filter
  %(prog)s news --from-date 01/05/2026 --to-date 14/05/2026 --filter "งบการเงิน"

  # Quick F45 earnings summary (no file download)
  %(prog)s quick --symbol BH --quarters 4

  # Quick F45 for all stocks filed on a date
  %(prog)s quick --from-date 14/05/2026 --to-date 14/05/2026

  # Quick F45 as JSON (for programmatic use)
  %(prog)s quick --symbol BH --quarters 2 --format json

  # Download BH financial statements (8 quarters)
  %(prog)s financial --symbol BH --quarters 8 --out ./output

  # Download ALL financial statements filed on a date
  %(prog)s financial --from-date 14/05/2026 --to-date 14/05/2026 --out ./output

  # List financial statements without downloading
  %(prog)s financial --symbol PTT --quarters 4 --list-only

  # Stock info — all sections
  %(prog)s stock --symbol BH

  # Stock info — specific sections only
  %(prog)s stock --symbol BH --sections price,shareholders

  # Historical trading with date range
  %(prog)s stock --symbol BH --sections historical --from-date 01/01/2026 --to-date 14/05/2026

  # Warrant info
  %(prog)s warrant --symbol ORI

  # List available reports (56-1 One Report, annual, ESG)
  %(prog)s report --symbol BH

  # List only 56-1 One Report
  %(prog)s report --symbol BH --type one

  # Download latest 56-1 One Report
  %(prog)s report --symbol BH --type one --out ./output

  # Download specific year
  %(prog)s report --symbol BH --type one --year 2024 --out ./output
""",
    )
    sub = parser.add_subparsers(dest="mode", required=True)

    # ── news ──
    p_news = sub.add_parser("news", help="Search SET company news")
    p_news.add_argument("--symbol", help="Stock symbol (omit for all stocks)")
    p_news.add_argument("--from-date", help="Start date DD/MM/YYYY")
    p_news.add_argument("--to-date", help="End date DD/MM/YYYY")
    p_news.add_argument("--keyword", help="Search keyword")
    p_news.add_argument("--filter", help="Regex to filter headlines (client-side)")
    p_news.add_argument("--limit", type=int, default=200, help="Max results (default: 200)")
    p_news.add_argument("--format", choices=["table", "json"], default="table",
                        help="Output format (default: table)")
    p_news.add_argument("--show-id", action="store_true", help="Show news IDs")

    # ── quick ──
    p_quick = sub.add_parser("quick", help="Quick view F45 earnings summaries (no download)")
    p_quick.add_argument("--symbol", help="Stock symbol (omit for all stocks)")
    p_quick.add_argument("--from-date", help="Start date DD/MM/YYYY")
    p_quick.add_argument("--to-date", help="End date DD/MM/YYYY")
    p_quick.add_argument("--quarters", type=int, default=4,
                         help="Number of quarters per symbol (default: 4, used with --symbol)")
    p_quick.add_argument("--limit", type=int, help="Max total results (used without --symbol)")
    p_quick.add_argument("--format", choices=["text", "json"], default="text",
                         help="Output format: text (default, for AI reading) or json")

    # ── financial ──
    p_fin = sub.add_parser("financial", help="Download financial statements")
    p_fin.add_argument("--symbol", help="Stock symbol (omit for all stocks)")
    p_fin.add_argument("--from-date", help="Start date DD/MM/YYYY")
    p_fin.add_argument("--to-date", help="End date DD/MM/YYYY")
    p_fin.add_argument("--quarters", type=int, default=8,
                       help="Number of quarters per symbol (default: 8, used with --symbol)")
    p_fin.add_argument("--limit", type=int, help="Max total downloads (used without --symbol)")
    p_fin.add_argument("--out", default=".", help="Output directory (default: .)")
    p_fin.add_argument("--list-only", action="store_true", help="List without downloading")
    p_fin.add_argument("--read-content", action="store_true",
                       help="After extracting ZIP, read and print content of xlsx/xls/docx files")

    # ── stock ──
    p_stock = sub.add_parser("stock", help="Stock information (price, shareholders, etc.)")
    p_stock.add_argument("--symbol", required=True, help="Stock symbol (required)")
    p_stock.add_argument("--sections",
                         help=("Comma-separated sections to fetch. "
                               "Available: price, historical, rights, shareholders, profile. "
                               "Default: all"))
    p_stock.add_argument("--from-date", help="Start date DD/MM/YYYY (for historical trading)")
    p_stock.add_argument("--to-date", help="End date DD/MM/YYYY (for historical trading)")

    # ── warrant ──
    p_warrant = sub.add_parser("warrant", help="Warrant information")
    p_warrant.add_argument("--symbol", required=True, help="Stock symbol (parent stock, e.g. ORI)")

    # ── report ──
    p_report = sub.add_parser("report", help="Annual reports, 56-1 One Report, ESG reports")
    p_report.add_argument("--symbol", required=True, help="Stock symbol (required)")
    p_report.add_argument("--type",
                          help=("Comma-separated report types. "
                                "Available: one (56-1 One Report), annual, form56, esg. "
                                "Default: all"))
    p_report.add_argument("--year", type=int, help="Filter by fiscal year (e.g. 2024)")
    p_report.add_argument("--limit", type=int, help="Max reports to download per type (default: 1)")
    p_report.add_argument("--out", help="Output directory (omit to list only)")
    p_report.add_argument("--list-only", action="store_true", help="List without downloading")

    args = parser.parse_args()

    if args.mode == "news":
        cmd_news(args)
    elif args.mode == "quick":
        cmd_quick(args)
    elif args.mode == "financial":
        cmd_financial(args)
    elif args.mode == "stock":
        cmd_stock(args)
    elif args.mode == "warrant":
        cmd_warrant(args)
    elif args.mode == "report":
        cmd_report(args)


if __name__ == "__main__":
    main()
